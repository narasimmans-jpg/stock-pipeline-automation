import pandas as pd
import numpy as np
from datetime import datetime
import warnings
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import logging
import sys
import requests
import json
import tempfile

warnings.filterwarnings('ignore')

# ============================================================
# CONFIGURATION — loaded from environment variables
# (Set these as GitHub Actions Secrets or in your .env file)
# ============================================================
EMAIL_USER      = os.getenv("EMAIL_USER",      "businesssupport@technosport.in")
EMAIL_PASSWORD  = os.getenv("EMAIL_PASSWORD",  "ctck cvix qafj dpoi")
RECIPIENT_EMAIL = os.getenv("RECIPIENT_EMAIL", "narasimman.s@technosport.in")

DROPBOX_REFRESH_TOKEN = os.getenv("DROPBOX_REFRESH_TOKEN", "Br_npYEnddsAAAAAAAAAAb1bLPOAv3SU8KPiqcbJYGgHY4R3Y3WdGP7BCcpe0F8h")
DROPBOX_APP_KEY       = os.getenv("DROPBOX_APP_KEY",       "eheunxwtckkpdwk")
DROPBOX_APP_SECRET    = os.getenv("DROPBOX_APP_SECRET",    "mo8qt53k93ov9cr")

# Dropbox paths (always use forward slashes, start with /)
DROPBOX_STOCK_PATH    = "/ODOO B2B REPORT/DATA FILE/stock_pipeline.xlsx"
DROPBOX_PIPELINE_PATH = "/ODOO B2B REPORT/DATA FILE/pipeline1.xlsx"
DROPBOX_OUTPUT_PATH   = "/ODOO B2B REPORT/DATA FILE/Stock_Pipeline_Analysis_Report.xlsx"

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler()
    ]
)

# Define styles to be removed from the output
BLANK_STYLES_TO_REMOVE = [
    "OR23", "OR55", "OR69", "OR56", "OR66", "OR24", "OR97", "OR03", "OR01A", "OR05",
    "OR07", "OR07B", "OR09", "OR09A", "OR12", "OR1A", "OR35", "OR51",
    "OR57", "OR96", "SWL01", "OR81", "CR86"
]

# ============================================================
# DROPBOX HELPERS
# ============================================================

def get_dropbox_access_token():
    """Get a fresh short-lived access token using the refresh token."""
    logging.info("🔑 Refreshing Dropbox access token...")
    url = "https://api.dropboxapi.com/oauth2/token"
    data = {
        "grant_type": "refresh_token",
        "refresh_token": DROPBOX_REFRESH_TOKEN,
        "client_id": DROPBOX_APP_KEY,
        "client_secret": DROPBOX_APP_SECRET,
    }
    response = requests.post(url, data=data)
    if response.status_code != 200:
        raise RuntimeError(f"Failed to refresh Dropbox token: {response.text}")
    access_token = response.json()["access_token"]
    logging.info("✅ Dropbox access token refreshed successfully")
    return access_token


def download_from_dropbox(access_token, dropbox_path):
    """Download a file from Dropbox and return a local temp file path."""
    logging.info(f"⬇️ Downloading from Dropbox: {dropbox_path}")
    url = "https://content.dropboxapi.com/2/files/download"
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Dropbox-API-Arg": json.dumps({"path": dropbox_path}),
    }
    response = requests.post(url, headers=headers)
    if response.status_code != 200:
        raise RuntimeError(f"Failed to download {dropbox_path}: {response.text}")

    suffix = os.path.splitext(dropbox_path)[1]
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    tmp.write(response.content)
    tmp.close()
    logging.info(f"✅ Downloaded to temp file: {tmp.name}")
    return tmp.name


def upload_to_dropbox(access_token, local_path, dropbox_path):
    """Upload a local file to Dropbox, overwriting if it exists."""
    logging.info(f"⬆️ Uploading to Dropbox: {dropbox_path}")
    url = "https://content.dropboxapi.com/2/files/upload"
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Dropbox-API-Arg": json.dumps({
            "path": dropbox_path,
            "mode": "overwrite",
            "autorename": False,
            "mute": False,
        }),
        "Content-Type": "application/octet-stream",
    }
    with open(local_path, "rb") as f:
        response = requests.post(url, headers=headers, data=f)
    if response.status_code not in (200, 201):
        raise RuntimeError(f"Failed to upload to Dropbox: {response.text}")
    logging.info("✅ Uploaded to Dropbox successfully")


# ============================================================
# ORIGINAL LOGIC (unchanged)
# ============================================================

def filter_blank_styles(df):
    """Remove specified blank styles from the dataframe"""
    if df.empty:
        return df

    initial_count = len(df)
    df['STYLE'] = df['STYLE'].astype(str).str.strip()
    styles_to_remove = set(style.strip() for style in BLANK_STYLES_TO_REMOVE)
    filtered_df = df[~df['STYLE'].isin(styles_to_remove)].copy()
    removed_count = initial_count - len(filtered_df)

    if removed_count > 0:
        logging.info(f"🗑️ Removed {removed_count} blank styles: {BLANK_STYLES_TO_REMOVE}")
        logging.info(f"📊 Remaining styles after filtering: {len(filtered_df)}")
        found_styles = set(df['STYLE']).intersection(styles_to_remove)
        if found_styles:
            logging.info(f"🔍 Specifically removed styles: {list(found_styles)}")
    else:
        logging.info("ℹ️ No blank styles found to remove")

    return filtered_df


def process_stock_data(stock_file_path):
    """Process stock data with robust column matching"""
    logging.info("📊 Processing Stock Data...")

    try:
        if not os.path.exists(stock_file_path):
            logging.error(f"❌ Stock file not found: {stock_file_path}")
            return pd.DataFrame()

        stock_df = pd.read_excel(stock_file_path)

        if stock_df.empty:
            logging.error("❌ Stock file is empty")
            return pd.DataFrame()

        original_columns = [str(col).strip() for col in stock_df.columns]
        logging.info(f"🔍 Columns found: {original_columns}")

        if 'Location' in stock_df.columns:
            initial_rows = len(stock_df)
            stock_df = stock_df[~stock_df['Location'].astype(str).str.contains('PACK/LFR', case=False, na=False)]
            filtered_rows = initial_rows - len(stock_df)
            logging.info(f"🗑️ Filtered out {filtered_rows} rows with Location='PACK/LFR'")
            logging.info(f"📊 Remaining rows: {len(stock_df)}")
        else:
            logging.warning("⚠️ 'Location' column not found - skipping PACK/LFR filter")

        if stock_df.empty:
            logging.error("❌ No data remaining after filtering PACK/LFR locations")
            return pd.DataFrame()

        column_mapping = {
            's': '06Y/S',
            'm': '08Y/M',
            'l': '10Y/L',
            'xl': '12Y/XL',
            '2xl': '14Y/2XL',
            '3xl': '3XL',
            '4xl': '4XL',
            '5xl': '5XL',
            '6uk': '06UK',
            '7uk': '07UK',
            '8uk': '08UK',
            '9uk': '09UK',
            '10uk': '10UK',
            '11uk': '11UK',
            'grand total': 'STOCK',
            'cat sales': 'CATEGORY'
        }

        rename_dict = {}
        for col in stock_df.columns:
            normalized_col = str(col).strip().lower()
            if normalized_col in column_mapping:
                rename_dict[col] = column_mapping[normalized_col]
                logging.info(f"✅ Renaming column: '{col}'➡️ '{column_mapping[normalized_col]}'")

        df = stock_df.rename(columns=rename_dict)

        if 'CATEGORY' not in df.columns:
            logging.warning("⚠️ 'CAT SALES' column not found after renaming.")
        else:
            logging.info("✅ CAT SALES column found and renamed to CATEGORY")

        size_mappings = [
            ('06Y', '06Y/S'),
            ('08Y', '08Y/M'),
            ('10Y', '10Y/L'),
            ('12Y', '12Y/XL'),
            ('14Y', '14Y/2XL')
        ]

        for old_col, new_col in size_mappings:
            if old_col in df.columns and new_col in df.columns:
                df[new_col] = df[old_col].fillna(0) + df[new_col].fillna(0)
                df = df.drop(columns=[old_col])
                logging.info(f"✅ Combined columns: {old_col} + {new_col}")
            elif old_col in df.columns:
                df = df.rename(columns={old_col: new_col})
                logging.info(f"✅ Renamed column: {old_col} ➡️ {new_col}")

        numeric_columns = ['06Y/S', '08Y/M', '10Y/L', '12Y/XL', '14Y/2XL',
                           '06UK', '07UK', '08UK', '09UK', '10UK', '11UK',
                           '3XL', '4XL', '5XL', 'STOCK']
        for col in numeric_columns:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

        group_cols = ['STYLE']
        if 'CATEGORY' in df.columns:
            group_cols.append('CATEGORY')

        if 'STYLE' not in df.columns:
            logging.error("❌ STYLE column not found in stock data")
            return pd.DataFrame()

        df = df.dropna(subset=['STYLE'])
        df = df[df['STYLE'].astype(str).str.strip() != '']

        if df.empty:
            logging.error("❌ No valid STYLE data found after cleaning")
            return pd.DataFrame()

        existing_numeric_cols = [col for col in numeric_columns if col in df.columns]
        df_grouped = df.groupby(group_cols)[existing_numeric_cols].sum().reset_index()
        df_grouped = df_grouped.sort_values('STYLE', ascending=True)

        logging.info(f"✅ Stock data processed successfully! Shape: {df_grouped.shape}")
        return df_grouped

    except Exception as e:
        logging.error(f"❌ Error processing stock data: {str(e)}", exc_info=True)
        return pd.DataFrame()


def process_pipeline_data(pipeline_file_path):
    """Process pipeline data according to requirements"""
    logging.info("📈 Processing Pipeline Data...")

    try:
        if not os.path.exists(pipeline_file_path):
            logging.warning(f"⚠️ Pipeline file not found: {pipeline_file_path}")
            return pd.DataFrame(columns=['STYLE', 'MONTH'])

        pipeline_df = pd.read_excel(pipeline_file_path)

        if pipeline_df.empty:
            logging.warning("⚠️ Pipeline file is empty")
            return pd.DataFrame(columns=['STYLE', 'MONTH'])

        original_columns = [str(col).strip() for col in pipeline_df.columns]
        logging.info(f"🔍 Pipeline columns found: {original_columns}")

        if 'STYLE NO' in pipeline_df.columns and 'STYLE' not in pipeline_df.columns:
            pipeline_df = pipeline_df.rename(columns={'STYLE NO': 'STYLE'})
            logging.info("✅ Renamed 'STYLE NO' to 'STYLE'")
        elif 'STYLE NO' in pipeline_df.columns and 'STYLE' in pipeline_df.columns:
            pipeline_df = pipeline_df.drop(columns=['STYLE NO'])

        if 'From' in pipeline_df.columns:
            pipeline_df = pipeline_df[pipeline_df['From'].str.contains('CATALOGUE SHEET', na=False, case=False)]
            logging.info(f"📊 Pipeline data after filtering: {len(pipeline_df)} rows")
        else:
            logging.warning("⚠️ 'From' column not found in pipeline data")

        if len(pipeline_df) == 0:
            logging.warning("⚠️ No rows found with 'CATALOGUE SHEET' in From column")
            return pd.DataFrame(columns=['STYLE', 'MONTH'])

        if 'STYLE' not in pipeline_df.columns:
            logging.error("❌ STYLE column not found")
            return pd.DataFrame(columns=['STYLE', 'MONTH'])

        month_col = None
        for col in pipeline_df.columns:
            if 'month' in col.lower():
                month_col = col
                break

        if not month_col:
            logging.warning("⚠️ No month column found")
            return pd.DataFrame(columns=['STYLE', 'MONTH'])

        oqty_col = None
        for col in pipeline_df.columns:
            col_lower = col.lower().strip()
            if 'o qty' in col_lower or 'oqty' in col_lower or 'o_qty' in col_lower:
                oqty_col = col
                break

        if not oqty_col:
            pipeline_df = pipeline_df[['STYLE', month_col]].copy()
            pipeline_df = pipeline_df.rename(columns={month_col: 'MONTH'})
        else:
            pipeline_df = pipeline_df[['STYLE', month_col, oqty_col]].copy()
            pipeline_df = pipeline_df.rename(columns={month_col: 'MONTH', oqty_col: 'O_QTY'})
            pipeline_df['O_QTY'] = pd.to_numeric(pipeline_df['O_QTY'], errors='coerce').fillna(0)

        pipeline_df = pipeline_df.dropna(subset=['STYLE', 'MONTH'])
        pipeline_df['STYLE'] = pipeline_df['STYLE'].astype(str).str.strip()
        pipeline_df['MONTH'] = pipeline_df['MONTH'].astype(str).str.strip()
        pipeline_df = pipeline_df[
            (pipeline_df['STYLE'] != '') &
            (pipeline_df['MONTH'] != '') &
            (pipeline_df['STYLE'].str.lower() != 'nan') &
            (pipeline_df['MONTH'].str.lower() != 'nan')
        ]

        if pipeline_df.empty:
            logging.warning("⚠️ No valid pipeline data after cleaning")
            return pd.DataFrame(columns=['STYLE', 'MONTH'])

        if 'O_QTY' in pipeline_df.columns:
            grouped_df = pipeline_df.groupby(['STYLE', 'MONTH'], as_index=False)['O_QTY'].sum()
            grouped_df['MONTH_WITH_QTY'] = grouped_df.apply(
                lambda row: f"{row['MONTH']}({int(row['O_QTY'])})" if row['O_QTY'] > 0 else row['MONTH'],
                axis=1
            )
            final_df = grouped_df.groupby('STYLE')['MONTH_WITH_QTY'].apply(
                lambda x: ', '.join(sorted(set(x)))
            ).reset_index()
            final_df.columns = ['STYLE', 'MONTH']
        else:
            final_df = pipeline_df.groupby('STYLE')['MONTH'].apply(
                lambda x: ', '.join(sorted(set(x.astype(str).str.strip())))
            ).reset_index()

        final_df = final_df.reset_index(drop=True)
        logging.info(f"✅ Pipeline data processed! Shape: {final_df.shape}")
        return final_df

    except Exception as e:
        logging.error(f"❌ Error processing pipeline data: {str(e)}", exc_info=True)
        return pd.DataFrame(columns=['STYLE', 'MONTH'])


def merge_and_finalize_data(stock_df, pipeline_df):
    """Merge stock and pipeline data and create final output"""
    logging.info("🔄 Merging and finalizing data...")

    try:
        if stock_df.empty:
            logging.error("❌ No stock data to process")
            return pd.DataFrame()

        stock_df = filter_blank_styles(stock_df)

        if stock_df.empty:
            logging.error("❌ No stock data remaining after filtering blank styles")
            return pd.DataFrame()

        if len(pipeline_df) == 0:
            logging.warning("⚠️ No pipeline data to merge")
            stock_df['MONTH'] = ''
        else:
            pipeline_df = filter_blank_styles(pipeline_df)
            stock_df['STYLE'] = stock_df['STYLE'].astype(str).str.strip()
            pipeline_df['STYLE'] = pipeline_df['STYLE'].astype(str).str.strip()
            try:
                stock_df = stock_df.merge(pipeline_df, on='STYLE', how='left')
                logging.info(f"✅ Merge completed. Rows: {len(stock_df)}")
            except Exception as e:
                logging.warning(f"⚠️ Error during merge: {e}")
                stock_df['MONTH'] = ''

        stock_df['MONTH'] = stock_df['MONTH'].fillna('')

        def create_series(style):
            if pd.isna(style) or style == '' or str(style).lower() == 'nan':
                return 'UNKNOWN-SERIES'
            style_str = str(style).strip()
            if len(style_str) == 0:
                return 'UNKNOWN-SERIES'
            if style_str.upper().startswith('O'):
                return 'OR-SERIES'
            else:
                return f"{style_str[0].upper()}-SERIES"

        stock_df['SERIES'] = stock_df['STYLE'].apply(create_series)

        column_order = ['SERIES', 'STYLE', '06Y/S', '08Y/M', '10Y/L', '12Y/XL', '14Y/2XL',
                        '06UK', '07UK', '08UK', '09UK', '10UK', '11UK',
                        '3XL', '4XL', '5XL', 'STOCK', 'MONTH']
        if 'CATEGORY' in stock_df.columns:
            column_order.insert(2, 'CATEGORY')

        existing_columns = [col for col in column_order if col in stock_df.columns]
        final_df = stock_df[existing_columns].copy()
        final_df = final_df.sort_values(['SERIES', 'STYLE']).reset_index(drop=True)

        logging.info(f"📊 Final data shape: {final_df.shape}")
        logging.info("✅ Data merged and finalized successfully!")
        return final_df

    except Exception as e:
        logging.error(f"❌ Error merging data: {str(e)}", exc_info=True)
        return pd.DataFrame()


def style_worksheet(ws, title, start_row=1, color='366092'):
    try:
        ws.cell(row=start_row, column=1, value=title)
        title_cell = ws.cell(row=start_row, column=1)
        title_cell.font = Font(size=16, bold=True, color='FFFFFF')
        title_cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        max_col = max(ws.max_column, 1)
        max_col = min(max_col, 12)
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=max_col)
        ws.row_dimensions[start_row].height = 30
    except Exception as e:
        logging.warning(f"⚠️ Error styling worksheet: {e}")


def apply_table_formatting(ws, df, start_row=3, highlight_totals=False):
    try:
        header_color = '4472C4'
        header_font_color = 'FFFFFF'
        alternating_color1 = 'F2F2F2'
        alternating_color2 = 'FFFFFF'
        totals_color = 'FFD966'
        border_color = 'D1D1D1'

        thin_border = Border(
            left=Side(style='thin', color=border_color),
            right=Side(style='thin', color=border_color),
            top=Side(style='thin', color=border_color),
            bottom=Side(style='thin', color=border_color)
        )

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=start_row, column=col_idx)
            cell.font = Font(bold=True, color=header_font_color, size=11)
            cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        for row_idx in range(start_row + 1, start_row + 1 + len(df)):
            is_totals_row = highlight_totals and (row_idx == start_row + len(df))
            if is_totals_row:
                fill_color = totals_color
            else:
                fill_color = alternating_color1 if (row_idx - start_row) % 2 == 0 else alternating_color2

            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is None:
                    cell.value = ""
                elif isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                cell.border = thin_border
                cell.font = Font(size=10, bold=is_totals_row)

    except Exception as e:
        logging.warning(f"⚠️ Error applying table formatting: {e}")


def adjust_column_widths(ws, df, start_row=3):
    try:
        for col_idx, col_name in enumerate(df.columns, 1):
            column_letter = get_column_letter(col_idx)
            max_length = len(str(col_name))
            sample_size = min(len(df), 100)
            for row_idx in range(sample_size):
                try:
                    cell_value = df.iloc[row_idx, col_idx - 1]
                    if pd.notna(cell_value):
                        cell_length = len(str(cell_value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    continue
            adjusted_width = min(max(max_length + 2, 10), 40)
            ws.column_dimensions[column_letter].width = adjusted_width
    except Exception as e:
        logging.warning(f"⚠️ Error adjusting column widths: {e}")


def create_excel_report(df, output_file):
    """Create a clean Excel report with well-formatted tables"""
    logging.info("📊 Creating Excel report...")

    try:
        if df.empty:
            logging.error("❌ Cannot create report: No data available")
            return False

        output_dir = os.path.dirname(output_file)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir)

        wb = Workbook()
        wb.remove(wb.active)

        data_ws = wb.create_sheet(title="Stock Pipeline Report")
        style_worksheet(data_ws, "STOCK AND PIPELINE ANALYSIS REPORT", color='366092')
        apply_table_formatting(data_ws, df, start_row=3)
        adjust_column_widths(data_ws, df, start_row=3)

        if 'CATEGORY' in df.columns:
            category_ws = wb.create_sheet(title="Category Summary")
            style_worksheet(category_ws, "CATEGORY SALES SUMMARY", color='70AD47')

            size_cols = ['06Y/S', '08Y/M', '10Y/L', '12Y/XL', '14Y/2XL',
                         '06UK', '07UK', '08UK', '09UK', '10UK', '11UK',
                         '3XL', '4XL', '5XL', 'STOCK']
            size_cols = [col for col in size_cols if col in df.columns]

            if size_cols:
                category_df = df.groupby('CATEGORY')[size_cols].sum().reset_index()
                totals = category_df.sum(numeric_only=True)
                totals['CATEGORY'] = 'TOTAL'
                category_df = pd.concat([category_df, pd.DataFrame([totals])], ignore_index=True)
                apply_table_formatting(category_ws, category_df, start_row=3, highlight_totals=True)
                adjust_column_widths(category_ws, category_df, start_row=3)

        summary_ws = wb.create_sheet(title="Executive Summary")
        style_worksheet(summary_ws, "EXECUTIVE SUMMARY", color='4472C4')

        total_stock = int(df['STOCK'].sum()) if 'STOCK' in df.columns else 0
        avg_stock   = int(df['STOCK'].mean()) if 'STOCK' in df.columns and len(df) > 0 else 0
        max_stock   = int(df['STOCK'].max()) if 'STOCK' in df.columns else 0
        min_stock   = int(df['STOCK'].min()) if 'STOCK' in df.columns else 0

        summary_data = [
            ["Metric", "Value"],
            ["Total Styles", len(df)],
            ["Total Stock", total_stock],
            ["Number of Series", df['SERIES'].nunique() if 'SERIES' in df.columns else 0],
            ["Average Stock per Style", avg_stock],
            ["Maximum Stock (Single Style)", max_stock],
            ["Minimum Stock (Single Style)", min_stock]
        ]

        for r_idx, row in enumerate(summary_data, 3):
            for c_idx, value in enumerate(row, 1):
                cell = summary_ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 3:
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )

        if 'SERIES' in df.columns and 'STOCK' in df.columns:
            summary_ws.cell(row=11, column=1, value="Series Breakdown").font = Font(bold=True, size=14)
            series_summary = df.groupby('SERIES').agg({'STYLE': 'count', 'STOCK': 'sum'}).reset_index()
            series_summary.columns = ['Series', 'Style Count', 'Total Stock']

            for c_idx, header in enumerate(["Series", "Style Count", "Total Stock"], 1):
                cell = summary_ws.cell(row=12, column=c_idx, value=header)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )

            for r_idx, (_, row) in enumerate(series_summary.iterrows(), 13):
                for c_idx, value in enumerate(row, 1):
                    cell = summary_ws.cell(row=r_idx, column=c_idx,
                                           value=int(value) if isinstance(value, (int, float)) else value)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
                    cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )

        for col_num in range(1, 4):
            summary_ws.column_dimensions[get_column_letter(col_num)].width = 25

        timestamp = datetime.now().strftime("Generated on %Y-%m-%d at %H:%M:%S")
        for sheet in wb:
            try:
                sheet.cell(row=sheet.max_row + 2, column=1, value=timestamp).font = Font(italic=True, color='808080')
            except:
                pass

        wb.save(output_file)
        logging.info(f"✅ Excel report created: {output_file}")
        return True

    except Exception as e:
        logging.error(f"❌ Error creating Excel report: {e}", exc_info=True)
        return False


def send_email(subject, body, to_emails, attachment_path=None):
    """Send email with optional attachment"""
    logging.info(f"✉️ Sending email...")

    sender_email    = EMAIL_USER
    sender_password = EMAIL_PASSWORD
    smtp_server     = "smtp.gmail.com"
    smtp_port       = 587

    if isinstance(to_emails, str):
        recipient_list = [e.strip() for e in to_emails.split(',') if e.strip()]
    else:
        recipient_list = to_emails
    recipient_list = [e for e in recipient_list if e and e.strip()]

    if not recipient_list:
        logging.error("❌ No valid email recipients found")
        return False

    try:
        msg = MIMEMultipart()
        msg['From']    = sender_email
        msg['To']      = ", ".join(recipient_list)
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'plain'))

        if attachment_path and os.path.exists(attachment_path):
            with open(attachment_path, 'rb') as attachment:
                part = MIMEApplication(attachment.read(), Name=os.path.basename(attachment_path))
            part['Content-Disposition'] = f'attachment; filename="{os.path.basename(attachment_path)}"'
            msg.attach(part)
        elif attachment_path:
            logging.warning(f"⚠️ Attachment not found: {attachment_path}")

        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(sender_email, sender_password)
            server.sendmail(sender_email, recipient_list, msg.as_string())

        logging.info(f"✅ Email sent to {len(recipient_list)} recipient(s)!")
        return True

    except Exception as e:
        logging.error(f"❌ Failed to send email: {str(e)}", exc_info=True)
        return False


# ============================================================
# MAIN RUN FUNCTION
# ============================================================

def run_report():
    """Main function — downloads files from Dropbox, processes, uploads result, sends email."""
    logging.info("🚀 Starting Stock and Pipeline Analysis")
    logging.info("=" * 60)

    tmp_stock    = None
    tmp_pipeline = None
    tmp_output   = None

    try:
        # 1. Get fresh Dropbox token
        access_token = get_dropbox_access_token()

        # 2. Download input files from Dropbox to temp files
        tmp_stock = download_from_dropbox(access_token, DROPBOX_STOCK_PATH)

        try:
            tmp_pipeline = download_from_dropbox(access_token, DROPBOX_PIPELINE_PATH)
        except Exception as e:
            logging.warning(f"⚠️ Could not download pipeline file: {e} — continuing without it")
            tmp_pipeline = None

        # 3. Create a temp path for the output Excel
        tmp_output_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        tmp_output_file.close()
        tmp_output = tmp_output_file.name

        # 4. Process data
        stock_data = process_stock_data(tmp_stock)
        if stock_data.empty:
            logging.error("❌ Aborting: Stock data processing failed")
            return False

        pipeline_data = process_pipeline_data(tmp_pipeline) if tmp_pipeline else pd.DataFrame(columns=['STYLE', 'MONTH'])
        final_data    = merge_and_finalize_data(stock_data, pipeline_data)

        if final_data.empty:
            logging.error("❌ Aborting: No final data to report")
            return False

        # 5. Create Excel report locally (temp file)
        success = create_excel_report(final_data, tmp_output)
        if not success:
            logging.error("❌ Failed to generate report")
            return False

        # 6. Upload result back to Dropbox
        upload_to_dropbox(access_token, tmp_output, DROPBOX_OUTPUT_PATH)

        # 7. Send email with the report as attachment
        recipient_emails = [e.strip() for e in RECIPIENT_EMAIL.split(',') if e.strip()]

        email_subject = "Daily Stock and Pipeline Analysis Report"
        email_body = """Hello Team,

Please find attached the daily Stock and Pipeline Analysis Report.

Key highlights:
- Total Styles: {styles:,}
- Total Stock: {stock:,}
- Number of Series: {series}
- Report generated at: {timestamp}

This is an automated report. Please contact the IT team if you have any issues.

Best regards,
Automated Reporting System""".format(
            styles=len(final_data),
            stock=int(final_data['STOCK'].sum()) if 'STOCK' in final_data.columns else 0,
            series=final_data['SERIES'].nunique() if 'SERIES' in final_data.columns else 0,
            timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        )

        send_success = send_email(
            subject=email_subject,
            body=email_body,
            to_emails=recipient_emails,
            attachment_path=tmp_output
        )

        if send_success:
            logging.info("✅ Report generated and sent successfully!")
            return True
        else:
            logging.error("❌ Report generated but failed to send email")
            return False

    except Exception as e:
        logging.error(f"❌ Error in report execution: {e}", exc_info=True)
        return False

    finally:
        # Clean up temp files
        for tmp_path in [tmp_stock, tmp_pipeline, tmp_output]:
            if tmp_path and os.path.exists(tmp_path):
                try:
                    os.remove(tmp_path)
                except:
                    pass
        logging.info("=" * 60)


if __name__ == "__main__":
    success = run_report()
    sys.exit(0 if success else 1)
